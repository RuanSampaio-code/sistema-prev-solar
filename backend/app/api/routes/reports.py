import io
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.core.database import get_db
from app.models.image import Image, ImageStatus
from app.models.result import Result
from app.models.user import User

router = APIRouter()

# ── cores do tema PrevSolar ──────────────────────────────────────────────────
_COLOR_HEADER_BG = "0F172A"   # slate-900
_COLOR_HEADER_FG = "FFFFFF"
_COLOR_PRIMARY   = "F59E0B"   # amber-400
_COLOR_ROW_ALT   = "F8FAFC"   # slate-50
_COLOR_TOTAL_BG  = "FEF3C7"   # amber-100
_COLOR_BORDER    = "E2E8F0"   # slate-200
_COLOR_GEO_BG    = "EFF6FF"   # blue-50

_THIN   = Side(style="thin",   color=_COLOR_BORDER)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _col(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _header_cell(ws, row: int, col: int, value: str, bg: str = _COLOR_HEADER_BG):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", size=10, bold=True, color=_COLOR_HEADER_FG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER
    return cell


def _build_resumo(ws, rows):
    """Aba 1 — resumo por imagem (uma linha por imagem processada)."""
    # Título
    ws.merge_cells("A1:G1")
    ws["A1"].value = "PrevSolar — Análise de Painéis Fotovoltaicos"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=_COLOR_PRIMARY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    ws["A2"].value = (
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        f"  |  {len(rows)} imagem(ns) processada(s)"
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="64748B")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    headers = ["Imagem", "Modelo", "Qtd. Painéis", "Área Total (m²)",
               "Potencial (kWh/mês)", "GSD (m/px)", "Processado em"]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 4, col, h)
    ws.row_dimensions[4].height = 22

    data_font = Font(name="Calibri", size=10)
    alt_fill  = PatternFill("solid", fgColor=_COLOR_ROW_ALT)
    total_panels, total_area, total_kwh = 0, 0.0, 0.0

    for i, (image, result) in enumerate(rows):
        r = i + 5
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        values = [
            image.original_name,
            (result.model_name or "—").upper(),
            result.panel_count,
            round(result.detected_area_m2, 2),
            round(result.estimated_kwh_month, 2),
            round(result.gsd_used_m_px, 6) if result.gsd_used_m_px else "—",
            result.processed_at.strftime("%d/%m/%Y %H:%M"),
        ]
        aligns = ["left", "center", "center", "center", "center", "center", "center"]
        for col, (val, align) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = _BORDER

        ws.cell(row=r, column=5).font = Font(name="Calibri", size=10, bold=True, color="059669")

        total_panels += result.panel_count
        total_area   += result.detected_area_m2
        total_kwh    += result.estimated_kwh_month

    # Totais
    if rows:
        tr = len(rows) + 5
        tf = PatternFill("solid", fgColor=_COLOR_TOTAL_BG)
        tf_font = Font(name="Calibri", size=10, bold=True)
        for col, val in enumerate(["TOTAL", "", total_panels, round(total_area, 2),
                                    round(total_kwh, 2), "", ""], 1):
            cell = ws.cell(row=tr, column=col, value=val)
            cell.fill = tf
            cell.font = tf_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    _col(ws, 1, 42); _col(ws, 2, 12); _col(ws, 3, 14)
    _col(ws, 4, 18); _col(ws, 5, 22); _col(ws, 6, 12); _col(ws, 7, 18)
    ws.freeze_panes = "A5"


def _build_paineis(ws, rows):
    """Aba 2 — painéis individuais (uma linha por painel detectado)."""
    ws.merge_cells("A1:L1")
    ws["A1"].value = "PrevSolar — Painéis Individuais"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=_COLOR_PRIMARY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6

    headers = [
        "#", "Imagem", "Modelo",
        "Área (m²)", "Geração (kWh/mês)", "Confiança (%)",
        "Centroide X (px)", "Centroide Y (px)",
        "Latitude", "Longitude", "Endereço",
    ]
    for col, h in enumerate(headers, 1):
        bg = "1E3A5F" if col >= 9 else _COLOR_HEADER_BG
        _header_cell(ws, 3, col, h, bg=bg)
    ws.row_dimensions[3].height = 22

    data_font  = Font(name="Calibri", size=9)
    geo_font   = Font(name="Calibri", size=9, color="1D4ED8")
    alt_fill   = PatternFill("solid", fgColor=_COLOR_ROW_ALT)
    geo_fill   = PatternFill("solid", fgColor=_COLOR_GEO_BG)
    panel_row  = 4

    for image, result in rows:
        panels = result.panels or []
        for panel in panels:
            r       = panel_row
            i_rel   = panel_row - 4
            base_fill = alt_fill if i_rel % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            lat     = panel.get("lat")
            lon     = panel.get("lon")
            endereco = panel.get("endereco", "")
            has_geo = lat is not None and lon is not None

            base_values = [
                panel.get("panel_id"),
                image.original_name,
                (result.model_name or "—").upper(),
                round(panel.get("area_m2", 0), 4),
                round(panel.get("kwh_month", 0), 2),
                round(panel.get("confidence_mean", 0) * 100, 1),
                panel.get("centroid_x"),
                panel.get("centroid_y"),
            ]
            aligns = ["center", "left", "center", "center", "center", "center", "center", "center"]

            for col, (val, align) in enumerate(zip(base_values, aligns), 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = base_fill
                cell.font = data_font
                cell.alignment = Alignment(horizontal=align, vertical="center")
                cell.border = _BORDER

            # kWh em verde
            ws.cell(row=r, column=5).font = Font(name="Calibri", size=9, bold=True, color="059669")

            # Lat / Lon / Endereço — azul se georreferenciado
            for col, val in enumerate([lat, lon, endereco], 9):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = geo_fill if has_geo else base_fill
                cell.font = geo_font if has_geo else data_font
                cell.alignment = Alignment(
                    horizontal="center" if col < 11 else "left",
                    vertical="center",
                    wrap_text=(col == 11),
                )
                cell.border = _BORDER

            panel_row += 1

    _col(ws, 1,  6);  _col(ws, 2,  38); _col(ws, 3, 10)
    _col(ws, 4, 13);  _col(ws, 5,  18); _col(ws, 6, 13)
    _col(ws, 7, 14);  _col(ws, 8,  14)
    _col(ws, 9, 14);  _col(ws, 10, 14); _col(ws, 11, 52)
    ws.freeze_panes = "A4"


@router.get("/xlsx")
def export_xlsx(
    image_id: int | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(Image, Result).join(Result).filter(Image.status == ImageStatus.done)
    if image_id:
        q = q.filter(Image.id == image_id)
    rows = q.order_by(Result.estimated_kwh_month.desc()).all()

    wb = Workbook()

    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    _build_resumo(ws_resumo, rows)

    ws_paineis = wb.create_sheet("Painéis Individuais")
    _build_paineis(ws_paineis, rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"prevsolar_relatorio_{image_id or 'completo'}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
